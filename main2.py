import io
import os
import cv2
import time
import base64
import random
import colorsys
import textwrap
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches

from openai import OpenAI
from dotenv import load_dotenv
from matplotlib import rcParams
from matplotlib.patches import Patch
from matplotlib.font_manager import FontProperties
from PIL import Image as PILImage
from googletrans import Translator
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage


load_dotenv()

COMIC = os.getenv('COMIC')
COMIC_ANIME_DIR = os.path.join(os.getenv('COMIC_ANIME_DIR'), COMIC, '1')
COMIC_DIR = os.path.join(os.getenv('COMIC_DIR'), COMIC)
DIALOGUE_DIR = os.path.join(os.getenv('DIALOGUE_DIR'), COMIC)
OBJECT_DIR = os.path.join(os.getenv('OBJECT_DIR'), COMIC)
GPT_O3_MODEL = os.getenv('GPT_O3_MODEL')
GPT_4O_MODEL = os.getenv('GPT_4O_MODEL')
GPT_KEY = os.getenv('GPT_KEY')
GEMINI_MODEL = os.getenv('GEMINI_MODEL')
# GEMINI_KEY = os.getenv('GEMINI_KEY')
GEMINI_KEYS_PATH = os.getenv('GEMINI_KEYS1_PATH')
GEMINI_INVALID_KEYS_PATH = os.getenv('GEMINI_INVALID_KEYS_PATH')
GEMINI_URL = os.getenv('GEMINI_URL')
PROMPT2_PATH = os.getenv('PROMPT2_PATH')
EXTENSION_DIR = os.getenv('EXTENSION_DIR')
OUTPUT_DIR = os.path.join(EXTENSION_DIR, COMIC, '1')

FONT_PATH = 'C:\\Windows\\Fonts\\SimHei.ttf'
FONT_PROP = FontProperties(fname=FONT_PATH)
rcParams['font.family'] = FONT_PROP.get_name()


def random_color():
    return tuple(random.randint(0, 255) for _ in range(3))


def random_colors(n):
    colors = []
    for idx in range(n):
        hue = idx / n
        r, g, b = colorsys.hsv_to_rgb(hue, 0.7, 1.0)
        colors.append((int(r * 255), int(g * 255), int(b * 255)))
    return colors


def wrap_text(text, width=20):
    return '\n'.join(textwrap.wrap(text, width=width))


def read_prompt(prompt_path):
    with open(prompt_path, 'r', encoding='utf-8') as file:
        content = file.read()
    return content


def image_to_base64(image_path):
    with open(image_path, 'rb') as img_file:
        return base64.b64encode(img_file.read()).decode('utf-8')


async def translate_text(text):
    translator = Translator()
    result = await translator.translate(text, dest='zh-CN')
    print(result)
    return result.text


def load_gemini_keys():
    with open(GEMINI_KEYS_PATH, 'r', encoding='utf-8') as f:
        return [line.strip() for line in f if line.strip()]


def load_gemini_invalid_keys():
    if os.path.exists(GEMINI_INVALID_KEYS_PATH):
        with open(GEMINI_INVALID_KEYS_PATH, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]
    return []


def save_gemini_invalid_key(key):
    with open(GEMINI_INVALID_KEYS_PATH, 'a') as f:
        f.write(f'{key}\n')


def get_response(model, model_key, prompt_content, base64_images, model_url=None):
    print('Model:', model)
    if model == GPT_4O_MODEL or model == GPT_O3_MODEL:
        client = OpenAI(api_key=model_key)
    else:
        client = OpenAI(base_url=model_url, api_key=model_key)
    content = [
        {
            'type': 'text',
            'text': prompt_content,
        }
    ]
    for base64_image in base64_images:
        content.append({
            'type': 'image_url',
            'image_url': {
                'url': f'data:image/jpeg;base64,{base64_image}'
            }
        })
    if model == GPT_O3_MODEL:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    'role': 'user',
                    'content': content
                }
            ],
            reasoning_effort='high'
        )
    elif model == GPT_4O_MODEL:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    'role': 'user',
                    'content': content
                }
            ],
            temperature=0
        )
    else:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    'role': 'user',
                    'content': content
                }
            ],
            reasoning_effort='high',
            temperature=0
        )
    return response.choices[0].message.content


def check_matching():
    for root, dirs, comic_anime_files in os.walk(COMIC_ANIME_DIR, topdown=True):
        for comic_anime_file in comic_anime_files:
            comic_anime_path = os.path.join(COMIC_ANIME_DIR, comic_anime_file)
            comic_anime_df = pd.read_csv(comic_anime_path)
            for index, row in comic_anime_df.iterrows():
                comic_block_id = row['Comic Block ID']
                if pd.isna(comic_block_id):
                    continue
                parts = comic_block_id.split('_')
                comic_id = parts[0]
                page_folder = f'page_{parts[2]}'
                image_path = None
                image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
                image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
                if os.path.exists(image_path_jpg):
                    image_path = image_path_jpg
                elif os.path.exists(image_path_png):
                    image_path = image_path_png
                if not image_path:
                    print(f'Image path is missing or invalid: {image_path_jpg, image_path_png}')
                    continue
                object_path = os.path.join(OBJECT_DIR, f'{comic_id}(已检查)', f'{page_folder}.csv')
                if not os.path.exists(object_path):
                    print(f'Object path is missing or invalid: {object_path}')
                    continue
                object_df = pd.read_csv(object_path)
                object_df = object_df[(object_df['image_name'] == os.path.basename(image_path))]
                if object_df.empty:
                    print(f'No object: {object_path, image_path}')
                    continue


def check_difference():
    label_summary_path = os.path.join(OBJECT_DIR, 'label_summary.csv')
    character_summary_path = os.path.join(DIALOGUE_DIR, 'character_summary.csv')
    label_summary_df = pd.read_csv(label_summary_path)
    character_summary_df = pd.read_csv(character_summary_path)
    label_names = set(label_summary_df['label_name'].astype(str))
    characters = set(character_summary_df['character'].astype(str))
    difference = characters - label_names
    print('Characters present in character_summary.csv but not in label_summary.csv:')
    for item in sorted(difference):
        print(item)


def show_bboxes(image_path, object_df, max_width=768, max_height=768):
    label_colors = {}
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f'Failed to load image: {image_path}')
    img_height, img_width = img.shape[:2]
    for _, row in object_df.iterrows():
        x = int(row['x'] * img_width)
        y = int(row['y'] * img_height)
        w = int(row['w'] * img_width)
        h = int(row['h'] * img_height)
        label = row['object']
        if label not in label_colors:
            label_colors[label] = random_color()
        color = label_colors[label]
        cv2.rectangle(img, (x, y), (x + w, y + h), color=color, thickness=20)
    scale = min(max_width / img_width, max_height / img_height, 1.0)
    if scale < 1.0:
        img_display = cv2.resize(img, (int(img_width * scale), int(img_height * scale)), interpolation=cv2.INTER_AREA)
    else:
        img_display = img
    cv2.imshow('Detected Objects', img_display)
    cv2.waitKey(0)
    cv2.destroyAllWindows()


def get_objects(comic_block_ids):
    objects = []
    for idx, comic_block_id in enumerate(comic_block_ids):
        parts = comic_block_id.split('_')
        comic_id = parts[0]
        page_folder = f'page_{parts[2]}'
        image_path = None
        image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
        image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
        if os.path.exists(image_path_jpg):
            image_path = image_path_jpg
        elif os.path.exists(image_path_png):
            image_path = image_path_png
        if not image_path:
            print(image_path)
            raise ValueError('Image path is missing or invalid.')
        object_path = os.path.join(OBJECT_DIR, f'{comic_id}(已检查)', f'{page_folder}.csv')
        if not os.path.exists(object_path):
            print(object_path)
            raise ValueError('Object path is missing or invalid.')
        object_df = pd.read_csv(object_path)
        object_df = object_df[(object_df['image_name'] == os.path.basename(image_path))]
        object_df['object'] = object_df['label_name']
        object_df['x'] = object_df['bbox_x'] / object_df['image_width']
        object_df['y'] = object_df['bbox_y'] / object_df['image_height']
        object_df['w'] = object_df['bbox_width'] / object_df['image_width']
        object_df['h'] = object_df['bbox_height'] / object_df['image_height']
        object_df = object_df[['object', 'x', 'y', 'w', 'h']]
        objects.append(object_df)
        # show_bboxes(image_path, object_df)
    return objects


def get_dialogues(comic_block_ids):
    dialogues = []
    for idx, comic_block_id in enumerate(comic_block_ids):
        parts = comic_block_id.split('_')
        comic_id = parts[0]
        page_folder = f'page_{parts[2]}'
        image_path = None
        image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
        image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
        if os.path.exists(image_path_jpg):
            image_path = image_path_jpg
        elif os.path.exists(image_path_png):
            image_path = image_path_png
        if not image_path:
            print(image_path_jpg, image_path_png)
            raise ValueError('Image path is missing or invalid.')
        dialogue_path = os.path.join(DIALOGUE_DIR, f'{comic_id}_updated(已检查).csv')
        if not os.path.exists(dialogue_path):
            print(dialogue_path)
            raise ValueError('Dialogue path is missing or invalid.')
        dialogue_df = pd.read_csv(dialogue_path)
        dialogue_df = dialogue_df[
            (dialogue_df['Page'] == page_folder) &
            (dialogue_df['Image'] == os.path.basename(image_path))
        ].fillna('旁白')
        dialogue_df = dialogue_df.rename(columns={'Character': 'object', 'Dialogue': 'dialogue'})
        dialogue_df = dialogue_df[['object', 'dialogue']]
        dialogues.append(dialogue_df)
    return dialogues


def get_base64_images(comic_block_ids):
    base64_images = []
    for idx, comic_block_id in enumerate(comic_block_ids):
        parts = comic_block_id.split('_')
        comic_id = parts[0]
        page_folder = f'page_{parts[2]}'
        image_path = None
        image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
        image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
        if os.path.exists(image_path_jpg):
            image_path = image_path_jpg
        elif os.path.exists(image_path_png):
            image_path = image_path_png
        if not image_path:
            print(image_path_jpg, image_path_png)
            raise ValueError('Image path is missing or invalid.')
        print(image_path)
        with PILImage.open(image_path) as img:
            width, height = img.size
            if width < height:
                scale = min(768 / width, 2000 / height)
            else:
                scale = min(768 / height, 2000 / width)
            new_width = int(width * scale)
            new_height = int(height * scale)
            resized_img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS).convert('RGB')
            buffer = io.BytesIO()
            resized_img.save(buffer, format='JPEG')
            base64_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
            base64_images.append(base64_str)
    return base64_images


def display_panels(comic_block_ids, objects, dialogues):
    num_panels = len(comic_block_ids)
    fig, axes = plt.subplots(2, num_panels, figsize=(3 * num_panels, 6), gridspec_kw={'height_ratios': [4, 1]})
    fig.subplots_adjust(bottom=0.2)
    if num_panels == 1:
        axes = axes.reshape(2, 1)
    for idx, (block_id, obj_df, dlg_df) in enumerate(zip(comic_block_ids, objects, dialogues)):
        parts = block_id.split('_')
        comic_id = parts[0]
        page_folder = f'page_{parts[2]}'
        image_path = None
        image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
        image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
        if os.path.exists(image_path_jpg):
            image_path = image_path_jpg
        elif os.path.exists(image_path_png):
            image_path = image_path_png
        if not image_path:
            print(image_path_jpg, image_path_png)
            raise ValueError('Image path is missing or invalid.')
        img = cv2.imread(image_path)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        max_width = 512
        scale = max_width / img.shape[1]
        img = cv2.resize(img, (int(img.shape[1] * scale), int(img.shape[0] * scale)))
        img_height, img_width = img.shape[:2]
        ax_img = axes[0, idx]
        ax_text = axes[1, idx]
        ax_img.imshow(img)
        ax_img.axis('off')
        panel_label = f'Panel {idx + 1}' if idx < num_panels - 1 else 'Last Panel'
        ax_img.set_title(f'{panel_label}: {comic_block_ids[idx]}')
        label_list = list(set(obj_df['object'].values))
        color_map = {label: color for label, color in zip(label_list, random_colors(len(label_list)))}
        for _, row in obj_df.iterrows():
            x = int(row['x'] * img_width)
            y = int(row['y'] * img_height)
            w = int(row['w'] * img_width)
            h = int(row['h'] * img_height)
            label = row['object']
            rgb = color_map[label]
            color = tuple(c / 255 for c in rgb)
            fill_rect = patches.Rectangle((x, y), w, h, linewidth=0, facecolor=color, alpha=0.25)
            border_rect = patches.Rectangle((x, y), w, h, linewidth=1, edgecolor=color, facecolor='none')
            ax_img.add_patch(fill_rect)
            ax_img.add_patch(border_rect)
            ax_img.text(x + w / 2, y + h / 2, label, color='black', fontsize=8, ha='center', va='center')
        legend_patches = [
            Patch(facecolor=(color_map[label][0] / 255.0, color_map[label][1] / 255.0, color_map[label][2] / 255.0),
                  label=label)
            for label in sorted(color_map.keys())
        ]
        ax_img.legend(handles=legend_patches, loc='upper center', bbox_to_anchor=(0.5, -0.15), fontsize=9,
                      ncol=3, frameon=False)
        dialogue_text = '\n\n'.join([wrap_text(f"{row['object']}: {row['dialogue']}") for _, row in dlg_df.iterrows()])
        ax_text.text(0, 1, f'\n{dialogue_text}', ha='left', va='center', wrap=True, fontsize=8)
        ax_text.axis('off')
    plt.tight_layout()
    plt.show()


def run(anime):
    output_path = os.path.join(OUTPUT_DIR, f'{anime}.pkl')
    if os.path.exists(output_path):
        df = pd.read_pickle(output_path)
        all_comic_blocks = df['comic_block_id'].tolist()
        responses = df['response'].tolist()
    else:
        df = pd.DataFrame(columns=['comic_block_id', 'response'])
        all_comic_blocks, responses = [], []
    comic_anime_path = os.path.join(COMIC_ANIME_DIR, f'{anime}.csv')
    comic_anime_df = pd.read_csv(comic_anime_path)
    for index, row in comic_anime_df.iterrows():
        current_comic_block_id = row['Comic Block ID']
        if pd.isna(current_comic_block_id):
            continue
        if current_comic_block_id in all_comic_blocks:
            continue
        all_comic_blocks.append(current_comic_block_id)
        start_idx = max(0, len(all_comic_blocks) - 6)
        pre_comic_block_ids = all_comic_blocks[start_idx:-1]
        comic_block_ids = pre_comic_block_ids
        comic_block_ids.append(current_comic_block_id)
        # print(comic_block_ids)
        objects = get_objects(comic_block_ids)
        # print(objects)
        dialogues = get_dialogues(comic_block_ids)
        # print(dialogues)
        # base64_images = get_base64_images(comic_block_ids)
        base64_images = get_base64_images([current_comic_block_id])
        num_panels = len(comic_block_ids)
        object_content = ''
        for idx, object_df in enumerate(objects):
            panel_title = (
                f'CSV #1 Panel {idx + 1}'
                if idx < num_panels - 1
                else 'CSV #1 Last Panel'
            )
            object_content += f'# {panel_title}\n{object_df.to_csv(index=False)}\n'
        object_content = f'```csv\n{object_content.rstrip()}\n```'
        dialogue_content = ''
        for idx, dialogue_df in enumerate(dialogues):
            panel_title = (
                f'CSV #2 Panel {idx + 1}'
                if idx < num_panels - 1
                else 'CSV #2 Last Panel'
            )
            dialogue_content += f'# {panel_title}\n{dialogue_df.to_csv(index=False)}\n'
        dialogue_content = f'```csv\n{dialogue_content.rstrip()}\n```'
        response_content = ''
        last_responses = responses[-(num_panels - 1):] if num_panels - 1 > 0 else []
        for idx, resp in enumerate(last_responses):
            panel_title = f'Description Panel {idx + 1}'
            response_content += f'# {panel_title}\n{resp}\n'
        response_content = f'```text\n{response_content.rstrip()}\n```'
        prompt_content = read_prompt(PROMPT2_PATH).format(COMIC, num_panels - 1, response_content)
        print(prompt_content)
        gemini_keys = load_gemini_keys()
        gemini_invalid_keys = load_gemini_invalid_keys()
        response = None
        for key in gemini_keys:
            if key in gemini_invalid_keys:
                continue
            print('Gemini Key:', key)
            while True:
                try:
                    response = get_response(GEMINI_MODEL, key, prompt_content, base64_images, GEMINI_URL)
                    break
                except Exception as e:
                    error_message = str(e)
                    if 'Error code: 429' in error_message:
                        print('Error 429:', e)
                        save_gemini_invalid_key(key)
                        break
                    elif 'Error code: 503' in error_message:
                        print('Error 503:', e)
                        time.sleep(30)
                        continue
                    else:
                        print('Error:', e)
                        break
            if response is not None:
                break
        if response is None:
            response = get_response(GPT_O3_MODEL, GPT_KEY, prompt_content, base64_images)
        print('Response:', response)
        # display_panels(comic_block_ids, objects, dialogues)
        df.loc[len(df)] = [current_comic_block_id, response]
        df.to_pickle(output_path)
        responses.append(response)
        print(f'[Saved] {current_comic_block_id} -> pickle ({len(df)} total)')


def show_output(anime):
    output_path = os.path.join(OUTPUT_DIR, f'{anime}.pkl')
    df = pd.read_pickle(output_path)
    max_image_size = 256
    char_per_line = 80
    line_height = 18
    temp_images = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comic Responses'
    ws.append(['comic_block_id', 'image', 'response'])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = max_image_size // 7
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 80
    for idx, row in df.iterrows():
        comic_block_id = row['comic_block_id']
        response = row['response'] if row['response'] else ''
        parts = comic_block_id.split('_')
        comic_id = parts[0]
        page_folder = f'page_{parts[2]}'
        image_path = None
        image_path_jpg = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.jpg')
        image_path_png = os.path.join(COMIC_DIR, comic_id, page_folder, f'{parts[3]}.png')
        if os.path.exists(image_path_jpg):
            image_path = image_path_jpg
        elif os.path.exists(image_path_png):
            image_path = image_path_png
        if not image_path:
            print(image_path_jpg, image_path_png)
            raise ValueError('Image path is missing or invalid.')
        ws.append([comic_block_id, '', response])
        image_height = 0
        if image_path:
            pil_img = PILImage.open(image_path)
            width, height = pil_img.size
            scale = min(max_image_size / width, max_image_size / height, 1.0)
            new_width = int(width * scale)
            new_height = int(height * scale)
            resized_path = os.path.join(EXTENSION_DIR, f'tmp_resized_{idx}.png')
            pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS).save(resized_path)
            image = XLImage(resized_path)
            temp_images.append(resized_path)
            image.anchor = f'B{idx + 2}'
            ws.add_image(image)
            image_height = new_height * 0.75
        for col_letter in ['C', 'D']:
            cell = ws[f'{col_letter}{idx + 2}']
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        line_count = len(response) // char_per_line + 1
        text_height = line_count * line_height
        row_height = max(image_height, text_height)
        ws.row_dimensions[idx + 2].height = row_height
    output_excel = os.path.splitext(output_path)[0] + '.xlsx'
    wb.save(output_excel)
    print(f'Saved Excel to: {output_excel}')
    for path in temp_images:
        if os.path.exists(path):
            os.remove(path)


def show_manual_output(anime):
    max_image_size = 256
    temp_images = []
    output_path = os.path.join(OUTPUT_DIR, f'{anime}_updated.xlsx')
    wb = load_workbook(output_path)
    ws = wb.active
    id_col = 'A'
    img_col = 'D'
    ws[f'{img_col}1'] = '即梦'
    ws.column_dimensions[img_col].width = 40
    for row in range(2, ws.max_row + 1):
        img_id = ws[f'{id_col}{row}'].value
        if not img_id:
            continue
        image_path = os.path.join(OUTPUT_DIR, anime, f'{img_id}.jpeg')
        if not os.path.exists(image_path):
            continue
        pil_img = PILImage.open(image_path)
        width, height = pil_img.size
        scale = min(max_image_size / width, max_image_size / height, 1.0)
        new_width = int(width * scale)
        new_height = int(height * scale)
        resized_path = os.path.join(EXTENSION_DIR, f'tmp_resized_{row}.jpeg')
        pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS).save(resized_path)
        temp_images.append(resized_path)
        img = XLImage(resized_path)
        img.anchor = f'{img_col}{row}'
        ws.add_image(img)
        current_height = ws.row_dimensions[row].height
        proposed_height = new_height * 0.75
        ws.row_dimensions[row].height = max(current_height, proposed_height)
    wb.save(output_path)
    print(f'Saved Excel with images to: {output_path}')
    for tmp in temp_images:
        if os.path.exists(tmp):
            os.remove(tmp)


if __name__ == '__main__':
    # print(load_gemini_keys())
    # check_matching()
    # check_difference()
    for i in range(22, 142):
        print(f'{i:03d}')
        run(f'{i:03d}')
        show_output(f'{i:03d}')
    # show_manual_output('144')
