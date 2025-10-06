import os
import glob

from dotenv import load_dotenv
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


load_dotenv()

COMIC = os.getenv('COMIC')
EXTENSION_DIR = os.path.join(os.getenv('EXTENSION_DIR'), 'Dragon Ball', '1')


def show_manual_output(anime):
    max_image_size = 256
    temp_images = []
    extension_path = os.path.join(EXTENSION_DIR, f'{anime}_updated.xlsx')
    wb = load_workbook(extension_path)
    ws = wb.active
    id_col = 'A'
    img_col = 'D'
    ws[f'{img_col}1'] = '即梦'
    ws.column_dimensions[img_col].width = 40
    for row in range(2, ws.max_row + 1):
        img_id = ws[f'{id_col}{row}'].value
        if not img_id:
            continue
        image_files = glob.glob(os.path.join(EXTENSION_DIR, anime, f'{img_id}.*'))
        image_path = next((f for f in image_files if f.lower().endswith(('.jpeg', '.jpg', '.png'))), None)
        if image_path is None:
            continue
        if not os.path.exists(image_path):
            continue
        pil_img = PILImage.open(image_path)
        width, height = pil_img.size
        scale = min(max_image_size / width, max_image_size / height, 1.0)
        new_width = int(width * scale)
        new_height = int(height * scale)
        resized_path = os.path.join(EXTENSION_DIR, f'tmp_resized_{row}.jpeg')
        pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
        if pil_img.mode == 'RGBA':
            pil_img = pil_img.convert('RGB')
        pil_img.save(resized_path)
        temp_images.append(resized_path)
        img = XLImage(resized_path)
        img.anchor = f'{img_col}{row}'
        ws.add_image(img)
        current_height = ws.row_dimensions[row].height
        proposed_height = new_height * 0.75
        ws.row_dimensions[row].height = max(current_height, proposed_height)
    wb.save(extension_path)
    print(f'Saved Excel with images to: {extension_path}')
    for tmp in temp_images:
        if os.path.exists(tmp):
            os.remove(tmp)


def show_undo(anime):
    extension_path = os.path.join(EXTENSION_DIR, f'{anime}_updated.xlsx')
    wb = load_workbook(extension_path)
    ws = wb.active
    first_col_values = [cell.value for cell in ws['A'][1:] if cell.value is not None]
    image_files = []
    for ext in ('*.jpg', '*.jpeg', '*.png'):
        image_files.extend(glob.glob(os.path.join(EXTENSION_DIR, anime, ext)))
    image_names = [os.path.splitext(os.path.basename(f))[0] for f in image_files]
    undo = [val for val in first_col_values if val not in image_names]
    print(undo)


if __name__ == '__main__':
    show_manual_output('040')
    # for i in range(1, 30):
    #     print(f'{i:03d}')
    #     show_undo(f'{i:03d}')
